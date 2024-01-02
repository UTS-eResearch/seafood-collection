from rocrate.rocrate import ROCrate # for creating the metadata files
from rocrate.model.preview import Preview # For creating the HTML representation of the ROCrates
from rocrate.model.entity import Entity
import os # for accessing the files
import openpyxl # For opening the excel
import csv # For saving the csv file
import matplotlib.pyplot as plt # For creating a diagram
from dateutil import parser # For handling ISO 8601 strings
import sys # For commandline usage
import re # For extracting years from filenames

# Plans:
# - Split into monthly csvs, from first of the month 00AM to last of the month 11:59PM - DONE!
# - Calculate max, min and avg values for readings over the month - DONE!
# - Create a diagram for temperature, depth and salinity = DONE!
# - Create an RO-Crate with those values - DONE!
# - Link up both the raw data and "working data" (decide whether to take readings from one or both?) - DONE!
# - Include a link in the ro-crate back to the website (or DOI) where this data can be obtained
# - Add the location - DONE!
# - Include author details - DONE!
# - Add notes around the data into RO-Crates for specific sensors and dates
# - Add local timezones

sub_folder = 'sensor-data'

# Globals from generate_data TODO: delete the unnecessary ones of these:
MONTHS = ["January","February","March","April","May","June","July","August","September","October","November","December"]
MONTH_NUMBERS = {"January":"01","February":"02","March":"03","April":"04","May":"05","June":"06","July":"07","August":"08","September":"09","October":"10","November":"11","December":"12"}
YEARS = ["2010","2011","2012","2013","2014","2015","2016","2017","2018","2019","2020"]
PLACES = [
    "Camden Haven","Clyde River","Georges River","Hawkesbury River","Hastings River","Manning River","Oyster Harbour","Pambula Lake", "Pindimar",
    "Pambula Lake","Port Stephens","Shoalhaven-Crookhaven River","Wagonga Inlet","Wallis Lake","Wonboyn Lake","Wapengo Lake", "Macleay River", "Merimbula Lake"
    ]
PLACE_CODES = {
    "Camden Haven":"CH","Clyde River":"CLY","Georges River":"GR","Hawkesbury River":"HR","Hastings River":"HS","Manning River":"Man","Pambula Lake - downstream":"PAM_02",
    "Pambula Lake - upstream":"PAM", "Pindimar":"PIN", "Port Stephens":"PS","Shoalhaven-Crookhaven River":"SH","Wagonga Inlet":"WAG","Wallis Lake":"WAL","Wonboyn Lake":"WON","Wapengo Lake":"WPG",
    "Oyster Harbour North":"WA_A1B","Oyster Harbour Central":"WA_A2","Oyster Harbour East":"WA_A3","Oyster Harbour East":"WA_A3B","Oyster Harbour South":"WA_A4",
    "Macleay River":"mncm2l301", "Hastings River":"mncm2l302", "Camden Haven":"mncm2l303", "Manning River":"mncm2l304", "Port Stephens":"mncm2l305","Georges River":"mncm2l306",
    "Shoalhaven-Crookhaven River":"mncm2l307", "Wagonga Inlet":"mncm2l308", "Wapengo Lake":"mncm2l309", "Merimbula Lake":"mncm2l30a", "Pambula Lake":"mncm2l30bc",
    "Wonboyn Lake":"mncm2l30d"
    }
PLACE_LOCATIONS = {
    "Camden Haven":"https://sws.geonames.org/8210175/","Clyde River":"https://sws.geonames.org/2171249/","Georges River":"https://sws.geonames.org/2205884/",
    "Hawkesbury River":"https://sws.geonames.org/2205605/","Hastings River":"https://sws.geonames.org/2163834/","Manning River":"https://sws.geonames.org/2158850/",
    "Pambula Lake - upstream":"https://sws.geonames.org/8594508/","Pambula Lake - downstream":"https://sws.geonames.org/8594508/",
    "Port Stephens":"https://sws.geonames.org/9409163/","Pindimar":"https://sws.geonames.org/2153153/","Shoalhaven-Crookhaven River":"https://sws.geonames.org/2149595/","Wagonga Inlet":"https://sws.geonames.org/2207090/",
    "Wallis Lake":"https://sws.geonames.org/8539070/","Wonboyn Lake":"https://sws.geonames.org/8210771/","Wapengo Lake":"https://sws.geonames.org/8594517/",
    "Oyster Harbour North":"https://sws.geonames.org/2064011/", "Oyster Harbour Central":"https://sws.geonames.org/2064011/", "Oyster Harbour East":"https://sws.geonames.org/2064011/",
    "Oyster Harbour South":"https://sws.geonames.org/2064011/", "Macleay River":"https://sws.geonames.org/2159179/", "Merimbula Lake":"https://sws.geonames.org/2158019/",
    "Pambula Lake":"https://sws.geonames.org/8594508/"
    }
LAT_LONG = {
    "Camden Haven":["-31.64478","152.82822"],"Clyde River":["-35.70093","150.13341"],"Georges River":["-34.02245","151.176"],"Hawkesbury River":["-33.5443","151.1365167"],
    "Hastings River":["-31.40406","152.89172"],"Manning River":["-31.89088","152.63981"],"Pambula Lake - downstream":["-36.96811903","149.884795"],
    "Pambula Lake - upstream":["-36.96811903","149.884795"],"Port Stephens":["-32.7196","152.06093"],"Shoalhaven-Crookhaven River":["-34.9118","150.74158"],
    "Wagonga Inlet":["-36.22161","150.07128"],"Wallis Lake":["-32.18268","152.47556"],"Wonboyn Lake":["-37.24121","149.92724"],"Wapengo Lake":["-36.60182","150.01678"],
    "Oyster Harbour North":["-34.964233", "117.955450"],"Oyster Harbour Central":["-34.973600", "117.959117"], "Oyster Harbour East":["-34.972117", "117.969433"],
    "Oyster Harbour South":["-34.986217","117.956850"], "Macleay River":["-30.86667","153.01667"], "Merimbula Lake":["-36.8985","149.8846"],
    "Pambula Lake":["-36.96811903","149.884795"], "Pindimar": ["-32.6828", "152.09807"]
    }
AUTHORS = [{
    "name":"Penelope Ajani",
    "email": "Penelope.Ajani@uts.edu.au",
    "FamilyName": "Ajani",
    "ORCID":"https://orcid.org/0000-0001-5364-9936",
    "givenName": "Penelope",
    "affiliation":{
        "@id": "https://ror.org/03f0f6041",
        "@type": "Organization",
        "Description": "The University of Technology Sydney is a public research university located in Sydney, Australia",
        "name": "University of Technology Sydney"
        }
    },
    {
    "name":"Shauna Murray",
    "email": "Shauna.Murray@uts.edu.au",
    "FamilyName": "Murray",
    "ORCID":"https://orcid.org/0000-0001-7096-1307",
    "givenName": "Shauna",
    "affiliation":{
        "@id": "https://ror.org/03f0f6041",
        "@type": "Organization",
        "Description": "The University of Technology Sydney is a public research university located in Sydney, Australia",
        "name": "University of Technology Sydney"
        }
    }]
# Globals:
NOTES_FILE = "2020-02-20 Summary working notes on data READ FIRST.xlsx"
RAW_SHEET = "raw" # Includes this phrase, prefixed by location
WORKING_SHEET = "Working" # Includes this phrase, prefixed by location
NEW_CSV_FILE_FOLDER = "split_data"

def get_file_name(raw_working,sensor,start_month,start_year,end_month,end_year):
    file_name = "out_" + sensor + "_" + start_month + start_year + "_" + end_month + end_year + "_" + raw_working
    return file_name

def select_sheet(name_part, workbook):
    sheets = workbook.get_sheet_names()
    for sheet_name in sheets:
        if name_part.lower() in sheet_name.lower():
            selected_sheet = sheet_name
    if selected_sheet is None:
        print("No sheet found with ",name_part, " in the name")
        return None
    else:
        return selected_sheet

def get_first_month_from_filename(filename):
    min_value = None
    month_name_length = 0
    for month_name in MONTHS:
        if month_name in filename:
            if min_value is None:
                min_value = filename.find(month_name)
                month_name_length = len(month_name)
            elif filename.find(month_name) < min_value:
                min_value = filename.find(month_name)
                month_name_length = len(month_name)
    if min_value:
        end_value = min_value + month_name_length
        return filename[min_value:end_value]
    else:
        return None



def get_first_year_from_filename(filename):
    matches = re.search("[0-9]{4}", filename)
    if matches:
        return matches.group()
    else:
        return None

def save_csv_file(sheet,header_row,start_row,end_row,file_name):
    print(f"csv file row length: {sheet.max_row}")
    if not os.path.exists(NEW_CSV_FILE_FOLDER):
        os.mkdir(NEW_CSV_FILE_FOLDER)
    with open(os.path.join(NEW_CSV_FILE_FOLDER,file_name + ".csv"),'w',newline='') as csv_file:
        fw = csv.writer(csv_file)
        fw.writerow(header_row)
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            fw.writerow(row_values)

def add_location(crate, geo_id, geonames_url, lat, long, name):
    place_properties = {"@type":"Place","name":name,"geo":{"@id":geo_id}}
    geo_properties = {"@type":"GeoCoordinates", "latitute":lat, "longitude":long, "name":name}
    place_entity = Entity(crate, identifier=geonames_url, properties=place_properties)
    geo_entity = Entity(crate, identifier=geo_id, properties=geo_properties)
    crate.add(geo_entity)
    crate.add(place_entity)
    update_root_dataset(crate,"contentLocation",{"@id":geonames_url})

def add_authors(crate, author_list):
    author = []
    affiliation = []
    person = []
    for author_dict in author_list:
        affiliation_dict = author_dict["affiliation"]
        id = author_dict["ORCID"]
        author.append({"@id":id})
        person.append({"@id":id, "@type":"Person","FamilyName":author_dict["FamilyName"],"givenName":author_dict["givenName"],"name":author_dict["name"],"affiliation":{"@id":author_dict["affiliation"]["@id"]}})
        affiliation.append(author_dict["affiliation"])
    for affiliation_entry in affiliation:
        affiliation_entity = Entity(crate, identifier=affiliation_entry["@id"],properties=affiliation_entry)
        crate.add(affiliation_entity)
    for person_entry in person:
        person_entity = Entity(crate, identifier=person_entry["@id"], properties=person_entry)
        crate.add(person_entity)
    update_root_dataset(crate, "author",author)



def add_year(crate, year):
    update_root_dataset(crate, "year", year)

def add_month(crate, month):
    update_root_dataset(crate,"month",month)

def update_root_dataset(crate, new_property_key, new_property_value):
    old_entity = None
    for entity in  list(crate.get_entities()):
        ent_props = entity.properties()
        if ent_props["@id"] == "./":
            old_entity = entity
    if old_entity:
        new_props = old_entity.properties()
        new_props[new_property_key] = new_property_value
        new_entity = Entity(crate, properties=new_props)
    else:
        print("Can't find the root dataset in this crate")
        return None

def add_quantitative_value(crate, id, value, unitCode, name, file_name):
    properties = {"@type": "QuantitativeValue", "name": name, "unitCode": unitCode, "value":value}
    qv = Entity(crate, identifier=id, properties=properties)
    #qv.type = "QuantititiveValue"
    crate.add(qv)
    update_file_entity(crate, file_name, name, {"@id":id})
    #### Need to link this added entity to the root dataset

def update_file_entity(crate, file_name, new_property_key, new_property_value):
    # Get the right entity first
    old_entity = None
    for entity in list(crate.get_entities()):
        ent_props = entity.properties()
        if file_name + ".csv" in ent_props.values(): # If the file name shows up in the properties of this entity
            old_entity = entity
    if old_entity:
        new_properties = old_entity.properties()
        new_properties[new_property_key] = new_property_value
        new_entity = Entity(crate, properties=new_properties)
    else:
        print("No file with that name in the crate: ",file_name + ".csv")
        return None


def package_data(crate,temperature,salinity,depth,location_name, month, year, file_name, data_entity):
    ## Plot the data, and set as crate image
    plot_three_datas(salinity, temperature, depth, file_name)
    crate.add_file(file_name + ".png")
    crate.image = file_name + ".png"



    ## Calculate the min, max and avg values, and add to the dataset in the crate
    if len(temperature) > 0:
        #print(temperature)
        avg_temp = round(sum(temperature)/len(temperature),4)
        min_temp = min(temperature)
        max_temp = max(temperature)
        # Add temperature values
        add_quantitative_value(crate, "avgTemperature", avg_temp, "CEL", "Average Temperature", file_name)
        add_quantitative_value(crate, "minTemperature", min_temp, "CEL", "Minimum Temperature", file_name)
        add_quantitative_value(crate, "maxTemperature", max_temp, "CEL", "Maximum Temperature", file_name)
    if len(salinity) > 0:
        avg_sal = round(sum(salinity)/len(salinity),4)
        min_sal = min(salinity)
        max_sal = max(salinity)
        # Add salinity values
        add_quantitative_value(crate, "avgSalinity", avg_sal, "NX", "Average Salinity", file_name)
        add_quantitative_value(crate, "minSalinity", min_sal, "NX", "Minimum Salinity", file_name)
        add_quantitative_value(crate, "maxSalinity", max_sal, "NX", "Maximum Salinity", file_name)
    if len(depth) > 0:
        avg_dep = round(sum(depth)/len(depth),4)
        min_dep = min(depth)
        max_dep = max(depth)
        # Add depth values
        add_quantitative_value(crate, "avgDepth", avg_dep, "MMT", "Average Depth", file_name)
        add_quantitative_value(crate, "minDepth", min_dep, "MMT", "Minimum Depth", file_name)
        add_quantitative_value(crate, "maxDepth", max_dep, "MMT", "Maximum Depth", file_name)

    #Add location
    add_location(crate, "_:b0", PLACE_LOCATIONS[location_name], LAT_LONG[location_name][0], LAT_LONG[location_name][1], location_name)

    # Add month
    #print("adding month")
    add_month(crate, month)
    # Add year
    add_year(crate, year)

    # Add the authors
    add_authors(crate,AUTHORS)

    # Need to have an existing Collection which has conformsTo
    # To handle Oni 2 indexer, add collection based on
    # Must have top-level license

    # list(crate.get_entities()) (returns a list of entities in the crate, can then do list[0].properties() to get properties of dataset)



# A method which takes in values and saves a (simple) plot of those values in the provided file_name
def plot_data(value_list, file_name):
    plt.plot(value_list)
    plt.savefig(file_name)
    plt.clf() # Clears the figure

def plot_three_datas(salinity, temperature, depth, file_name):
    figure, (salinity_ax, temperature_ax, depth_ax) = plt.subplots(3,1, sharex=False)
    print(f"length of salinity: {len(salinity)}")
    print(f"length of temperature: {len(temperature)}")
    print(f"length of depth: {len(depth)}")
    salinity_ax.plot(salinity, color='yellow')
    temperature_ax.plot(temperature, color='red')
    depth_ax.plot(depth, color='blue')
    salinity_ax.set_title('Salinity', fontsize=10)
    salinity_ax.axes.get_xaxis().set_ticklabels([]) # Hides Axis Label
    temperature_ax.set_title('Temperature', fontsize=10)
    temperature_ax.axes.get_xaxis().set_ticklabels([]) # Hides Axis Label
    depth_ax.set_title('Depth',fontsize=10)
    depth_ax.axes.get_xaxis().set_ticklabels([]) # Hides Axis Label
    figure.subplots_adjust(hspace=0.3) # Allows for enough space between the subplots to see titles
    plt.savefig(file_name + ".png")
    plt.clf()

def create_monthly_ro_crate(temperature,salinity,depth,file_name, location_name, month, year, csv_files, filepath=NEW_CSV_FILE_FOLDER):
    crate = ROCrate(gen_preview=False) # Don't generate a preview, since we'll do that using https://github.com/UTS-eResearch/ro-crate-excel
    crate.name = 'Sensor readings from ' + location_name + " in " + month + " " + str(year)
    for csv_file in csv_files:
        file_entity = crate.add_file(os.path.join(filepath,csv_file + ".csv")) # Adds the file reference to the crate, and will cause it to be saved next to it when written to disk
    print("start packaging data")
    package_data(crate,temperature,salinity,depth,location_name, month, year, csv_files[-1], file_entity) # Should use working csv file name (csv_files[-1]), since it needs the full .csv reference
    print("finish packaging data")
    crate_filepath = os.path.join(sub_folder, file_name)
    crate.write_crate(crate_filepath)
    os.system('xlro -j ' + crate_filepath)
    # Delete the image file now that it's in the crate - TODO This removes all instances of the image?
    os.remove(csv_files[-1] + ".png")
    return file_entity
    #preview = Preview(crate) # Create a preview of the crate
    #preview.write(file_name) # Write that preview out to the same folder

def get_location_code_by_name(filename):
    print("Looking for location name in: ", filename)
    for name, code in PLACE_CODES.items():
        if name.lower() in filename.lower():
            return code
    return None

def get_location_code_in_filename(filename):
    print("Looking for location code in: ", filename)
    for name, code in PLACE_CODES.items():
        if code in filename:
            return code
    return None

def get_location_name(location_code):
    for name, code in PLACE_CODES.items():
        if code in location_code:
            return name
    for name, code in PLACE_CODES.items():
        if code.lower() in location_code.lower():
            return name
    return None

def process_sheet(sheet, raw_status, location_code):
    value_dict = {}
    header_row = True
    headers = []
    first_row= None
    last_row = None
    start_month = None
    start_year = None
    depth = []
    temperature = []
    salinity = []
    # Need to record first row, and then last row of the month, in order to export data
    for row in sheet.rows:
        # Deal with the header row
        if header_row:
            print(row)
            header_row = False
            for cell in row:
                headers.append(cell.value)
        else:
            # If this is our first row, set the start values
            if start_month is None:
                start_row = row[0].row
                if row[headers.index("MeasurementTime")].value is not None:
                    date = parser.parse(row[headers.index("MeasurementTime")].value)
                    if (date is not None):
                        start_month = date.strftime("%B")
                        start_year = date.year
            # Otherwise we want to check if we've moved to a different month yet
            else:
                # Double check that there is a  value then parse it
                if row[headers.index("MeasurementTime")].value is not None:
                    date = parser.parse(row[headers.index("MeasurementTime")].value)
                    # If it is a separate month, save the previous rows
                    if date.strftime("%B") != start_month:
                        end_month = date.strftime("%B")
                        end_year = date.year
                        end_row = row[0].row - 1
                        print("Saving rows: ",start_row, " to ", end_row)
                        file_name = get_file_name(raw_status, location_code + "_01", str(start_month), str(start_year), str(end_month), str(end_year))
                        print("in file: ", file_name + ".csv")
                        save_csv_file(sheet,headers,start_row,end_row, file_name)
                        value_dict[file_name] = {"salinity":salinity, "temperature":temperature,"depth":depth,"start_month":start_month,"start_year":start_year}


                        # Reset values for the next month
                        start_month = date.strftime("%B")
                        start_year = date.year
                        start_row = row[0].row
                        temperature = []
                        depth = []
                        salinity = []

            # Record the various values for avg, mins and max
            # print(f'Type: |{row[headers.index("Type")].value}|')
            # print(f'CurrentValue: |{row[headers.index("CurrentValue")].value}|')
            # print(f'type of value: {type(row[headers.index("Type")].value)}')
            if row[headers.index("Type")].value == "temperature":
                temperature.append(row[headers.index("CurrentValue")].value)
                # print(f'found temperature, temperature length: {len(temperature)}')
            if row[headers.index("Type")].value == "water depth":
                depth.append(row[headers.index("CurrentValue")].value)
                # print(f'found water depth, depth length: {len(depth)}')
            if row[headers.index("Type")].value == "salinity":
                salinity.append(row[headers.index("CurrentValue")].value)
                # print(f'found salinity, salinity: {len(salinity)}')
            # print()
    # Handles final rows
    end_row = sheet.max_row
    print("Saving rows: ",start_row, " to ", end_row)
    csv_file_name = get_file_name(raw_status,location_code + "_01", str(start_month), str(start_year), str(end_month), str(end_year))
    print("in file: ", csv_file_name + ".csv")
    save_csv_file(sheet,headers,start_row,end_row,csv_file_name)
    value_dict[csv_file_name] = {"salinity":salinity, "temperature":temperature,"depth":depth,"start_month":start_month,"start_year":start_year}
    #create_monthly_ro_crate(temperature,salinity,depth, file_name, location_name, start_month, start_year)
    return value_dict

def process_api_csv(csv_rows, raw_status="RAW", location_code=None, file_name=None):
    value_dict = {}
    header_row = True
    headers = []
    start_row = 0
    end_row = 0
    first_row= None
    last_row = None
    start_month = None # Need to set this from filename as fallback
    start_year = None # Need to set this from filename as fallback
    depth = []
    temperature = []
    salinity = []
    # Need to record first row, and then last row of the month, in order to export data
    for index, row in enumerate(csv_rows):
        # Deal with the header row
        if header_row:
            #print(row)
            header_row = False
            for cell in row:
                headers.append(cell)
        else:
            # If this is our first row, set the start values
            if start_month is None:
                start_row = 1
                if row[headers.index("MeasurementTime")] is not None:
                    date = parser.parse(row[headers.index("MeasurementTime")])
                    if (date is not None):
                        start_month = date.strftime("%B")
                        start_year = date.year
            # Otherwise we want to check if we've moved to a different month yet
            else:
                # Double check that there is a  value then parse it
                if row[headers.index("MeasurementTime")] is not None:
                    date = parser.parse(row[headers.index("MeasurementTime")])
                    # If it is a separate month, save the previous rows
                    if date.strftime("%B") != start_month:
                        end_month = date.strftime("%B")
                        end_year = date.year
                        end_row = index
                        print("Saving rows: ",start_row, " to ", end_row)
                        if not file_name:
                            file_name = get_file_name(raw_status, location_code + "_01", str(start_month), str(start_year), str(end_month), str(end_year))
                            print("in file: ", file_name + ".csv")
                            save_csv_file(sheet,headers,start_row,end_row, file_name)
                        value_dict[file_name] = {"salinity":salinity, "temperature":temperature,"depth":depth,"start_month":start_month,"start_year":start_year}

            # Record the various values for avg, mins and max
            if row[headers.index("Type")] == "temperature":
                temperature.append(float(row[headers.index("CurrentValue")]))
                #print("Found temp", row[headers.index("Type")],float(row[headers.index("CurrentValue")]))
            elif row[headers.index("Type")] == "depth":
                depth.append(float(row[headers.index("CurrentValue")]))
                #print("Found depth", row[headers.index("Type")],float(row[headers.index("CurrentValue")]))
            elif row[headers.index("Type")] == "salinity":
                salinity.append(float(row[headers.index("CurrentValue")]))
                #print("Found salinity", row[headers.index("Type")],float(row[headers.index("CurrentValue")]))

    # Handles final rows
    #print("Salnity",salinity)
    end_row = len(csv_rows) - 1
    print("Processing rows: ",start_row, " to ", end_row)
    csv_file_name = file_name
    print("from file: ", csv_file_name)
    #save_csv_file(sheet,headers,start_row,end_row,csv_file_name)
    if start_month is None: # No values to get month from:
        start_month = get_first_month_from_filename(file_name)
    if start_year is None:
        start_year = get_first_year_from_filename(file_name)

    value_dict[csv_file_name] = {"salinity":salinity, "temperature":temperature,"depth":depth,"start_month":start_month,"start_year":start_year}
    #create_monthly_ro_crate(temperature,salinity,depth, file_name, location_name, start_month, start_year)
    return value_dict

def working_from_raw_filename(raw_filename):
    if "RAW" in raw_filename:
        working_filename = raw_filename.replace("RAW","WORKING")
    else:
        working_filename = raw_filename.replace("raw","working")
    return working_filename

def process_data(raw_sheet, working_sheet, location_code, location_name):
    raw_dict = process_sheet(raw_sheet, "RAW", location_code)
    working_dict = process_sheet(working_sheet, "WORKING", location_code)
    for raw_key, raw_values in raw_dict.items():
        plain_filename = raw_key.removesuffix("_RAW")
        working_key = working_from_raw_filename(raw_key)
        if working_key in working_dict:
            print("working key in working dict")
            working_values = working_dict[working_key]
            create_monthly_ro_crate(working_values["temperature"],working_values["salinity"],working_values["depth"],plain_filename,location_name,working_values["start_month"],working_values["start_year"],[raw_key, working_key]) # Use the working values when create something
        else:
            print("working key not in working dict")
            create_monthly_ro_crate(raw_values["temperature"],raw_values["salinity"],raw_values["depth"],plain_filename,location_name,raw_values["start_month"],raw_values["start_year"],[raw_key]) # Use raw values if no working ones available
    #for csv_file_pair in csv_files:
    #    create_monthly_ro_crate(temperature,salinity,depth, file_name, location_name, start_month, start_year, csv_file_pair)

def process_api_data(csv_file_path, csv_file_name):
    location_code = get_location_code_in_filename(csv_file_name)
    print("Location Code: ",location_code)
    location_name = get_location_name(location_code)
    csv_rows = []
    with open(os.path.join(csv_file_path,csv_file_name),'r') as file:
        reader = csv.reader(file, delimiter=',')
        for row in reader:
            csv_rows.append(row)
    value_dict = process_api_csv(csv_rows,file_name=csv_file_name)

    for key, values in value_dict.items():
        #print("key",key)
        #print("val",values)
        #print("salinity",values["salinity"])
        plain_filename = key.removesuffix(".csv")
        # for k,v in values.items():
        #     print(k)

        create_monthly_ro_crate(values["temperature"],values["salinity"],values["depth"],plain_filename,location_name,values["start_month"],values["start_year"],[plain_filename],csv_file_path)

def load_api_data_from_folder(filepath):
    directory = os.path.join(os.getcwd(),filepath)
    print("Loading files in: ", directory)
    for file in os.listdir(directory):
        if os.path.isfile(os.path.join(directory,file)) and ".csv" in file:
            location_name = get_location_name(file)
            process_api_data(directory, file)

def load_estuary_data(filepath):
    location_code = get_location_code_by_name(filepath)
    location_name = get_location_name(location_code) # Why have one line, when you can have two
    print("Loading data from: ",os.path.join(os.getcwd(),filepath))
    workbook = openpyxl.load_workbook(filepath, read_only=True) # Read-only mode so that it doesn't take up too much memory

    # process RAW and WORKING data
    raw_sheet_name = select_sheet(RAW_SHEET, workbook)
    working_sheet_name = select_sheet(WORKING_SHEET, workbook)
    if raw_sheet_name is None or working_sheet_name is None:
        return None
    print("Using sheets: ", raw_sheet_name, " and ", working_sheet_name)
    raw_sheet = workbook[raw_sheet_name]
    working_sheet = workbook[working_sheet_name]
    print(f"location_code: {location_code}, location_name: {location_name}")
    process_data(raw_sheet, working_sheet, location_code, location_name)

    #print("Min/Max temperature: ",min(temperature), " / ", max(temperature))
    workbook.close() # Need to close since read-only uses "lazy loading"

if __name__ == "__main__":
    # TODO: Split option between API data and exported data
    # TODO: List all files in a path, and then process them
    if len(sys.argv) == 4:
        first_option = sys.argv[1]
        second_option = sys.argv[2]
        third_option = sys.argv[3]
        if second_option == "--folder":
            if first_option == "--api":
                load_api_data_from_folder(third_option)
            elif first_option == "--downloaded":
                load_estuary_data(third_option)
        elif second_option == "--file":
            if first_option == "--api":
                process_api_data(os.getcwd(),third_option)
            elif first_option == "--downloaded":
                print("I don't have a function for that yet...")
        else:
            print("Make sure you specify either --file or --folder in the second position")
    else:
        print("Wrong commandline amount of options present:",len(sys.argv), " should be 5")
        print("Usage: python3 open_data.py <--api/--downloaded>, <--file/--folder>, <file/folder path>")
