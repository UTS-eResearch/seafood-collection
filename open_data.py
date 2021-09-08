from rocrate.rocrate import ROCrate # for creating the metadata files
from rocrate.model.preview import Preview # For creating the HTML representation of the ROCrates
from rocrate.model.entity import Entity
import os # for accessing the files
import openpyxl # For opening the excel
import csv # For saving the csv file
import matplotlib.pyplot as plt # For creating a diagram
from dateutil import parser # For handling ISO 8601 strings

# Plans:
# - Split into monthly csvs, from first of the month 00AM to last of the month 11:59PM - DONE!
# - Calculate max, min and avg values for readings over the month - DONE!
# - Create a diagram for temperature, depth and salinity = DONE!
# - Create an RO-Crate with those values - DONE!
# - Link up both the raw data and "working data" (decide whether to take readings from one or both?)
# - Include a link in the ro-crate back to the website (or DOI) where this data can be obtained
# - Add notes around the data into RO-Crates for specific sensors and dates


# Globals from generate_data TODO: delete the unnecessary ones of these:
MONTHS = ["January","February","March","April","May","June","July","August","September","October","November","December"]
MONTH_NUMBERS = {"January":"01","February":"02","March":"03","April":"04","May":"05","June":"06","July":"07","August":"08","September":"09","October":"10","November":"11","December":"12"}
YEARS = ["2010","2011","2012","2013","2014","2015","2016","2017","2018","2019","2020"]
PLACES = ["Camden Haven","Clyde River","Georges River","Hawkesbury River","Hastings River","Manning River","Pambula Lake","Pambula Lake","Port Stephens","Shoalhaven Crookhaven Rivers","Wagonga Inlet","Wallis Lake","Wonboyn Lake","Wapengo Lake"]
PLACE_CODES = {"Camden Haven":"CH","Clyde River":"CLY","Georges River":"GR","Hawkesbury River":"HR","Hastings River":"HR","Manning River":"Man","Pambula Lake":"PAM","Port Stephens":"PS","Shoalhaven Crookhaven Rivers":"SH","Wagonga Inlet":"WAG","Wallis Lake":"WAL","Wonboyn Lake":"WON","Wapengo Lake":"WPG"}
PLACE_LOCATIONS = {"Camden Haven":"https://sws.geonames.org/8210175/","Clyde River":"https://sws.geonames.org/2171249/","Georges River":"https://sws.geonames.org/2205884/","Hawkesbury River":"https://sws.geonames.org/2205605/","Hastings River":"https://sws.geonames.org/2163834/","Manning River":"https://sws.geonames.org/2158850/","Pambula Lake":"https://sws.geonames.org/8594508/","Port Stephens":"https://sws.geonames.org/9409163/","Shoalhaven Crookhaven Rivers":"https://sws.geonames.org/2149595/","Wagonga Inlet":"https://sws.geonames.org/2207090/","Wallis Lake":"https://sws.geonames.org/8539070/","Wonboyn Lake":"https://sws.geonames.org/8210771/","Wapengo Lake":"https://sws.geonames.org/8594517/"}
LAT_LONG = {"Camden Haven":["-31.64478","152.82822"],"Clyde River":["-35.70093","150.13341"],"Georges River":["-34.02245","151.176"],"Hawkesbury River":["-33.5443","151.1365167"],"Hastings River":["-31.40406","152.89172"],"Manning River":["-31.89088","152.63981"],"Pambula Lake":["-36.96811903","149.884795"],"Port Stephens":["-32.7196","152.06093"],"Shoalhaven Crookhaven Rivers":["-34.9118","150.74158"],"Wagonga Inlet":["-36.22161","150.07128"],"Wallis Lake":["-32.18268","152.47556"],"Wonboyn Lake":["-37.24121","149.92724"],"Wapengo Lake":["-36.60182","150.01678"]}

# Globals:
NOTES_FILE = "2020-02-20 Summary working notes on data READ FIRST.xlsx"
RAW_SHEET = "raw data" # Includes this phrase, prefixed by location
WORKING_SHEET = "Working data" # Includes this phrase, prefixed by location
NEW_CSV_FILE_FOLDER = "split_data"

def get_file_name(raw_working,sensor,start_month,start_year,end_month,end_year):
    file_name = "out_" + sensor + "_" + start_month + start_year + "_" + end_month + end_year + "_" + raw_working
    return file_name

def select_sheet(name_part, workbook):
    sheets = workbook.get_sheet_names()
    for sheet_name in sheets:
        if name_part.lower() in sheet_name.lower():
            selected_sheet = sheet_name
    if selected_sheet is None:
        print("No sheet found with ",name_part, " in the name")
        return None
    else:
        return selected_sheet

def save_csv_file(sheet,header_row,start_row,end_row,file_name):

    if not os.path.exists(NEW_CSV_FILE_FOLDER):
        os.mkdir(NEW_CSV_FILE_FOLDER)
    with open(os.path.join(NEW_CSV_FILE_FOLDER,file_name + ".csv"),'w',newline='') as csv_file:
        fw = csv.writer(csv_file)
        fw.writerow(header_row)
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            fw.writerow(row_values)

def add_quantitative_value(crate, id, value, unitCode, name, file_name):
    properties = {"@type": "QuantitativeValue", "name": name, "unitCode": unitCode, "value":value}
    qv = Entity(crate, identifier=id, properties=properties)
    #qv.type = "QuantititiveValue"
    crate.add(qv)
    update_file_entity(crate, file_name, name, {"@id":id})
    #### Need to link this added entity to the root dataset

def update_file_entity(crate, file_name, new_property_key, new_property_value):
    # Get the right entity first
    for entity in list(crate.get_entities()):
        ent_props = entity.properties()
        if file_name + ".csv" in ent_props.values(): # If the file name shows up in the properties of this entity
            old_entity = entity
    if old_entity:
        new_properties = old_entity.properties()
        new_properties[new_property_key] = new_property_value
        new_entity = Entity(crate, properties=new_properties)
    else:
        print("No file with that name in the crate :(")
        return None


def package_data(crate,temperature,salinity,depth,file_name, data_entity):
    ## Plot the data, and set as crate image
    plot_three_datas(salinity, temperature, depth, file_name)
    crate.add_file(file_name + ".png")
    crate.image = file_name + ".png"
    ## Calculate the min, max and avg values, and add to the dataset in the crate
    avg_temp = round(sum(temperature)/len(temperature),4)
    min_temp = min(temperature)
    max_temp = max(temperature)
    avg_sal = round(sum(salinity)/len(salinity),4)
    min_sal = min(salinity)
    max_sal = max(salinity)
    avg_dep = round(sum(depth)/len(depth),4)
    min_dep = min(depth)
    max_dep = max(depth)
    # Add temperature values
    add_quantitative_value(crate, "avgTemperature", avg_temp, "CEL", "Average Temperature", file_name)
    add_quantitative_value(crate, "minTemperature", min_temp, "CEL", "Minimum Temperature", file_name)
    add_quantitative_value(crate, "maxTemperature", max_temp, "CEL", "Maximum Temperature", file_name)
    # Add salinity values
    add_quantitative_value(crate, "avgSalinity", avg_sal, "NX", "Average Salinity", file_name)
    add_quantitative_value(crate, "minSalinity", min_sal, "NX", "Minimum Salinity", file_name)
    add_quantitative_value(crate, "maxSalinity", max_sal, "NX", "Maximum Salinity", file_name)
    # Add depth values
    add_quantitative_value(crate, "avgDepth", avg_dep, "MMT", "Average Depth", file_name)
    add_quantitative_value(crate, "minDepth", min_dep, "MMT", "Minimum Depth", file_name)
    add_quantitative_value(crate, "maxDepth", max_dep, "MMT", "Maximum Depth", file_name)



    # list(crate.get_entities()) (returns a list of entities in the crate, can then do list[0].properties() to get properties of dataset)



# A method which takes in values and saves a (simple) plot of those values in the provided file_name
def plot_data(value_list, file_name):
    plt.plot(value_list)
    plt.savefig(file_name)
    plt.clf() # Clears the figure

def plot_three_datas(salinity, temperature, depth, file_name):
    figure, (salinity_ax, temperature_ax, depth_ax) = plt.subplots(3,1, sharex=False)
    salinity_ax.plot(salinity, color='yellow')
    temperature_ax.plot(temperature, color='red')
    depth_ax.plot(depth, color='blue')
    salinity_ax.set_title('Salinity', fontsize=10)
    salinity_ax.axes.get_xaxis().set_ticklabels([]) # Hides Axis Label
    temperature_ax.set_title('Temperature', fontsize=10)
    temperature_ax.axes.get_xaxis().set_ticklabels([]) # Hides Axis Label
    depth_ax.set_title('Depth',fontsize=10)
    depth_ax.axes.get_xaxis().set_ticklabels([]) # Hides Axis Label
    figure.subplots_adjust(hspace=0.3) # Allows for enough space between the subplots to see titles
    plt.savefig(file_name + ".png")
    plt.clf()

def create_monthly_ro_crate(temperature,salinity,depth,file_name, location_name, month, year):
    crate = ROCrate(gen_preview=False) # Don't generate a preview, since we'll do that using https://github.com/UTS-eResearch/ro-crate-excel
    crate.name = 'Sensor readings from ' + location_name + " in " + month + " " + str(year)
    file_entity = crate.add_file(os.path.join(NEW_CSV_FILE_FOLDER,file_name + ".csv")) # Adds the file reference to the crate, and will cause it to be saved next to it when written to disk
    package_data(crate,temperature,salinity,depth,file_name, file_entity)
    crate.write_crate(file_name)
    os.system('xlro -j ' + file_name)
    return file_entity
    #preview = Preview(crate) # Create a preview of the crate
    #preview.write(file_name) # Write that preview out to the same folder

def get_location_code(filename):
    for name, code in PLACE_CODES.items():
        if name.lower() in filename.lower():
            return code
    return None

def get_location_name(location_code):
    for name, code in PLACE_CODES.items():
        if code.lower() in location_code.lower():
            return name
    return None

def load_estuary_data(filepath):
    location_code = get_location_code(filepath)
    location_name = get_location_name(location_code) # Why have one line, when you can have two
    print("Loading data from: ",os.path.join(os.getcwd(),filepath))
    workbook = openpyxl.load_workbook(filepath, read_only=True) # Read-only mode so that it doesn't take up too much memory
    active_sheet_name = select_sheet(RAW_SHEET, workbook)
    if active_sheet_name is None:
        return None
    print("Using sheet: ", active_sheet_name)
    active_sheet = workbook[active_sheet_name]

    header_row = True
    headers = []
    #### Split here?
    first_row= None
    last_row = None
    start_month = None
    start_year = None
    depth = []
    temperature = []
    salinity = []
    # Need to record first row, and then last row of the month, in order to export data
    for row in active_sheet.rows:
        # Deal with the header row
        if header_row:
            print(row)
            header_row = False
            for cell in row:
                headers.append(cell.value)
        else:
            # If this is our first row, set the start values
            if start_month is None:
                start_row = row[0].row
                date = parser.parse(row[headers.index("MeasurementTime")].value)
                if (date is not None):
                    start_month = date.strftime("%B")
                    start_year = date.year
            # Otherwise we want to check if we've moved to a different month yet
            else:
                date = parser.parse(row[headers.index("MeasurementTime")].value)
                # If it is a separate month, save the previous rows
                if date.strftime("%B") != start_month:
                    end_month = date.strftime("%B")
                    end_year = date.year
                    end_row = row[0].row - 1
                    print("Saving rows: ",start_row, " to ", end_row)
                    file_name = get_file_name("RAW","CH_01", str(start_month), str(start_year), str(end_month), str(end_year))
                    print("in file: ", file_name + ".csv")
                    save_csv_file(active_sheet,headers,start_row,end_row, file_name)


                    create_monthly_ro_crate(temperature,salinity,depth, file_name, location_name, start_month, start_year)
                    # Reset values for the next month
                    start_month = date.strftime("%B")
                    start_year = date.year
                    start_row = row[0].row
                    tempature = []
                    depth = []
                    salinity = []

            # Record the various values for avg, mins and max
            if row[headers.index("Type")].value == "temperature":
                temperature.append(row[headers.index("CurrentValue")].value)
            elif row[headers.index("Type")].value == "water depth":
                depth.append(row[headers.index("CurrentValue")].value)
            elif row[headers.index("Type")].value == "salinity":
                salinity.append(row[headers.index("CurrentValue")].value)
    # Handles final rows
    end_row = active_sheet.max_row
    print("Saving rows: ",start_row, " to ", end_row)
    csv_file_name = get_file_name("RAW","CH_01", str(start_month), str(start_year), str(end_month), str(end_year))
    print("in file: ", csv_file_name)
    save_csv_file(active_sheet,headers,start_row,end_row,csv_file_name)



    print("Min/Max temperature: ",min(temperature), " / ", max(temperature))
    workbook.close() # Need to close since read-only uses "lazy loading"
